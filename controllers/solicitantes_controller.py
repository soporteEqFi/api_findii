from __future__ import annotations

from flask import request, jsonify, make_response
from models.solicitantes_model import SolicitantesModel
from models.ubicaciones_model import UbicacionesModel
from models.actividad_economica_model import ActividadEconomicaModel
from models.informacion_financiera_model import InformacionFinancieraModel
from models.referencias_model import ReferenciasModel
from models.solicitudes_model import SolicitudesModel
from utils.debug_helpers import (
    log_request_details, log_validation_results,
    log_data_to_save, log_operation_result, log_response, log_error
)
from utils.email.sent_email import enviar_email_registro_completo
import json
import os
import uuid
import unicodedata
import io
from datetime import datetime, timezone, timedelta
from werkzeug.utils import secure_filename
from data.supabase_conn import supabase
from models.documentos_model import DocumentosModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class SolicitantesController:
    def __init__(self):
        self.model = SolicitantesModel()
        self.ubicaciones_model = UbicacionesModel()
        self.actividad_model = ActividadEconomicaModel()
        self.financiera_model = InformacionFinancieraModel()
        self.referencias_model = ReferenciasModel()
        self.solicitudes_model = SolicitudesModel()
        self.documentos_model = DocumentosModel()

    def _empresa_id(self) -> int:
        empresa_id = request.headers.get("X-Empresa-Id") or request.args.get("empresa_id")
        if not empresa_id:
            raise ValueError("empresa_id es requerido")
        try:
            return int(empresa_id)
        except Exception as exc:
            raise ValueError("empresa_id debe ser entero") from exc

    def create(self):
        log_request_details("CREAR SOLICITANTE", "solicitantes")

        try:
            empresa_id = self._empresa_id()
            body = request.get_json(silent=True) or {}

            print(f"\n📋 EMPRESA ID: {empresa_id}")

            # Campos requeridos
            required_fields = [
                "nombres", "primer_apellido", "segundo_apellido",
                "tipo_identificacion", "numero_documento",
                "fecha_nacimiento", "genero", "correo"
            ]

            # Validar campos requeridos
            if not log_validation_results(required_fields, body):
                raise ValueError("Faltan campos requeridos")

            # Preparar datos para guardar
            datos_a_guardar = {
                "empresa_id": empresa_id,
                "nombres": body["nombres"],
                "primer_apellido": body["primer_apellido"],
                "segundo_apellido": body["segundo_apellido"],
                "tipo_identificacion": body["tipo_identificacion"],
                "numero_documento": body["numero_documento"],
                "fecha_nacimiento": body["fecha_nacimiento"],
                "genero": body["genero"],
                "correo": body["correo"],
                "info_extra": body.get("info_extra"),
            }

            log_data_to_save(datos_a_guardar)

            # Crear en BD
            data = self.model.create(**datos_a_guardar)

            log_operation_result(data, "SOLICITANTE CREADO")

            response_data = {"ok": True, "data": data}
            log_response(response_data)

            return jsonify(response_data), 201
        except ValueError as ve:
            log_error(ve, "ERROR DE VALIDACIÓN")
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            log_error(ex, "ERROR INESPERADO")
            return jsonify({"ok": False, "error": str(ex)}), 500

    def get_one(self, id: int):
        try:
            empresa_id = self._empresa_id()
            data = self.model.get_by_id(id=id, empresa_id=empresa_id)
            if not data:
                return jsonify({"ok": False, "error": "No encontrado"}), 404
            return jsonify({"ok": True, "data": data})
        except ValueError as ve:
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            return jsonify({"ok": False, "error": str(ex)}), 500

    def list(self):
        try:
            empresa_id = self._empresa_id()
            limit = int(request.args.get("limit", 50))
            offset = int(request.args.get("offset", 0))
            data = self.model.list(empresa_id=empresa_id, limit=limit, offset=offset)
            return jsonify({"ok": True, "data": data})
        except ValueError as ve:
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            return jsonify({"ok": False, "error": str(ex)}), 500

    def descargar_ventas_excel(self):
        """Exportar todos los solicitantes a Excel (.xlsx) con columnas configurables"""
        try:
            empresa_id = self._empresa_id()
            print(f"\n📥 DESCARGANDO VENTAS EXCEL - Empresa ID: {empresa_id}")
            
            # Obtener todos los datos sin límite
            data = self.model.list(empresa_id=empresa_id, limit=10000, offset=0)
            print(f"   📊 Total de registros: {len(data)}")
            
            # CONFIGURACIÓN DE COLUMNAS - Orden exacto del frontend
            # Formato: ("Nombre en Excel", función_transformación)
            columnas_config = [
                ("Nombres", lambda item: f"{item.get('nombres', '')} {item.get('primer_apellido', '')} {item.get('segundo_apellido', '')}".strip()),
                ("Identificación", lambda item: item.get("tipo_identificacion", "")),
                ("Fecha", lambda item: self._extraer_fecha(item.get("created_at", ""))),
                ("Hora", lambda item: self._extraer_hora(item.get("created_at", ""))),
                ("Número Documento", lambda item: item.get("numero_documento", "")),
                ("Ciudad Residencia", lambda item: item.get("ciudad_residencia", "")),
                ("Correo", lambda item: item.get("correo", "")),
                ("Celular", lambda item: item.get("celular", "")),
                ("Tipo Crédito", lambda item: item.get("tipo_credito", "")),
                ("Banco", lambda item: item.get("banco_nombre", "")),
                ("Estado", lambda item: item.get("estado_solicitud", "")),
                ("Creado por", lambda item: item.get("created_by_user_name", "")),
                ("Supervisor", lambda item: item.get("created_by_supervisor_name", "")),
                # AGREGAR MÁS COLUMNAS AQUÍ:
                # ("Nueva Columna", lambda item: item.get("nuevo_campo", "")),
            ]
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas"
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir encabezados
            headers = [col[0] for col in columnas_config]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            print(f"   📋 Columnas exportadas: {', '.join(headers)}")
            
            # Escribir datos
            for row_idx, item in enumerate(data, start=2):
                for col_idx, (_, extractor) in enumerate(columnas_config, start=1):
                    try:
                        value = extractor(item)
                        # Convertir None a string vacío
                        if value is None:
                            value = ""
                        # Asegurar que sea string
                        value = str(value) if value else ""
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    except Exception as e:
                        print(f"   ⚠️ Error extrayendo valor en fila {row_idx}, col {col_idx}: {e}")
                        ws.cell(row=row_idx, column=col_idx, value="")
            
            # Ajustar ancho de columnas basado en el contenido real
            for col_idx, header in enumerate(headers, start=1):
                column_letter = ws.cell(row=1, column=col_idx).column_letter
                
                # Calcular el ancho máximo basado en el contenido
                max_length = len(str(header))  # Empezar con el largo del encabezado
                
                # Revisar todas las filas para encontrar el valor más largo
                for row_idx in range(2, len(data) + 2):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        cell_length = len(str(cell_value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                # Ajustar el ancho con un margen adicional (máximo 50 para evitar columnas muy anchas)
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar en memoria con manejo de errores
            output = io.BytesIO()
            try:
                wb.save(output)
                output.seek(0)
                excel_data = output.getvalue()
                print(f"   💾 Archivo Excel guardado en memoria: {len(excel_data)} bytes")
            except Exception as save_error:
                print(f"   ❌ Error al guardar Excel: {save_error}")
                raise
            finally:
                output.close()
            
            # Preparar respuesta
            response = make_response(excel_data)
            filename = f"ventas_{empresa_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Headers optimizados para descarga de Excel
            response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-Length"] = str(len(excel_data))
            response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
            
            print(f"   ✅ Excel generado exitosamente con {len(columnas_config)} columnas y {len(data)} registros")
            return response
            
        except ValueError as ve:
            log_error(ve, "ERROR DE VALIDACIÓN EN DESCARGA EXCEL")
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            log_error(ex, "ERROR INESPERADO EN DESCARGA EXCEL")
            return jsonify({"ok": False, "error": str(ex)}), 500
    
    def _extraer_fecha(self, created_at: str) -> str:
        """Extrae la fecha en formato DD/MM/YYYY del timestamp, convertido a zona horaria de Colombia (UTC-5)"""
        if not created_at:
            return ""
        try:
            # Parsear el timestamp UTC
            dt_utc = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            
            # Convertir a zona horaria de Colombia (UTC-5)
            colombia_tz = timezone(timedelta(hours=-5))
            dt_colombia = dt_utc.astimezone(colombia_tz)
            
            return dt_colombia.strftime("%d/%m/%Y")
        except Exception as e:
            try:
                return created_at.split("T")[0] if "T" in str(created_at) else str(created_at)
            except:
                return ""
    
    def _extraer_hora(self, created_at: str) -> str:
        """Extrae la hora en formato 12 horas (hh:mm:ss AM/PM) del timestamp, convertido a zona horaria de Colombia (UTC-5)"""
        if not created_at:
            return ""
        try:
            # Parsear el timestamp UTC
            dt_utc = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            
            # Convertir a zona horaria de Colombia (UTC-5)
            colombia_tz = timezone(timedelta(hours=-5))
            dt_colombia = dt_utc.astimezone(colombia_tz)
            
            # Formato 12 horas con AM/PM
            return dt_colombia.strftime("%I:%M:%S %p")
        except Exception as e:
            try:
                if "T" in str(created_at):
                    return created_at.split("T")[1].split(".")[0]
                return ""
            except:
                return ""

    def update(self, id: int):
        try:
            empresa_id = self._empresa_id()
            body = request.get_json(silent=True) or {}
            if not body:
                raise ValueError("Body requerido para actualizar")

            data = self.model.update(id=id, empresa_id=empresa_id, updates=body)
            return jsonify({"ok": True, "data": data})
        except ValueError as ve:
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            return jsonify({"ok": False, "error": str(ex)}), 500

    def delete(self, id: int):
        try:
            empresa_id = self._empresa_id()
            deleted = self.model.delete(id=id, empresa_id=empresa_id)
            return jsonify({"ok": True, "deleted": deleted})
        except ValueError as ve:
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            return jsonify({"ok": False, "error": str(ex)}), 500

    def crear_registro_completo(self):
        """Crear un registro completo: solicitante + ubicaciones + actividad económica + info financiera + referencias + solicitudes"""
        log_request_details("CREAR REGISTRO COMPLETO", "solicitante + todas las entidades relacionadas")

        try:
            empresa_id = self._empresa_id()

            # Soportar JSON puro y multipart/form-data con 'payload' + archivos
            content_type = request.content_type or ""
            if "multipart/form-data" in content_type:
                print(f"   📨 Content-Type recibido: {content_type}")
                try:
                    print(f"   📝 Form keys: {list(request.form.keys())}")
                    print(f"   📁 Files keys: {list(request.files.keys())}")
                except Exception as _e:
                    print(f"   ⚠️ No se pudieron listar keys de form/files: {_e}")
                raw_payload = request.form.get("payload")
                body = json.loads(raw_payload) if raw_payload else {}
                # Recolectar posibles campos de archivos enviados
                files_list = []
                for key in [
                    "documentos", "documentos[]", "files", "files[]", "file"
                ]:
                    if key in request.files:
                        # getlist también funciona para una sola entrada
                        files_list.extend(request.files.getlist(key))
                print(f"   📎 Archivos recibidos (multipart): {len(files_list)}")
            else:
                print(f"   📨 Content-Type recibido: {content_type} (sin multipart, no hay archivos)")
                body = request.get_json(silent=True) or {}
                files_list = []

            print(f"\n📋 EMPRESA ID: {empresa_id}")
            print(f"\n📦 DATOS RECIBIDOS:")
            print(f"   Claves principales: {list(body.keys())}")

            # Extraer datos de cada entidad
            datos_solicitante = body.get("solicitante", {})
            print(f"   🔍 Body completo: {body}")
            print(f"   🔍 Claves en body: {list(body.keys())}")
            print(f"   🔍 datos_solicitante: {datos_solicitante}")
            print(f"   🔍 Tipo datos_solicitante: {type(datos_solicitante)}")

            # Manejar ubicaciones (puede ser objeto o lista)
            ubicacion_obj = body.get("ubicacion", {})
            ubicaciones_list = body.get("ubicaciones", [])
            datos_ubicaciones = [ubicacion_obj] if ubicacion_obj else ubicaciones_list

            datos_actividad = body.get("actividad_economica", {})
            datos_financiera = body.get("informacion_financiera", {})

            # Manejar referencias (puede ser objeto o lista)
            referencia_obj = body.get("referencia", {})
            referencias_list = body.get("referencias", [])
            datos_referencias = [referencia_obj] if referencia_obj else referencias_list

            # Manejar solicitudes (puede ser objeto o lista)
            solicitud_obj = body.get("solicitud", {})
            solicitudes_list = body.get("solicitudes", [])
            datos_solicitudes = [solicitud_obj] if solicitud_obj else solicitudes_list

            print(f"\n🔍 DATOS EXTRAÍDOS:")
            print(f"   Solicitante: {len(datos_solicitante)} campos")
            print(f"   Ubicaciones: {len(datos_ubicaciones)} registros")
            print(f"   Actividad económica: {'✅ Presente' if datos_actividad else '❌ Vacío'}")
            print(f"   Info financiera: {'✅ Presente' if datos_financiera else '❌ Vacío'}")
            print(f"   Referencias: {len(datos_referencias)} registros")
            print(f"   Solicitudes: {len(datos_solicitudes)} registros")

            # Debug: mostrar qué se encontró
            if ubicacion_obj:
                print(f"   📍 Ubicación encontrada como objeto")
            if referencia_obj:
                print(f"   👥 Referencia encontrada como objeto")
            if solicitud_obj:
                print(f"   📋 Solicitud encontrada como objeto")

            # 1. CREAR SOLICITANTE
            print(f"\n1️⃣ CREANDO SOLICITANTE...")
            if not datos_solicitante or len(datos_solicitante) == 0:
                raise ValueError("Datos del solicitante son requeridos")

            datos_solicitante["empresa_id"] = empresa_id
            print(f"   Datos a guardar: {datos_solicitante}")

            solicitante_creado = self.model.create(**datos_solicitante)
            solicitante_id = solicitante_creado["id"]
            print(f"   ✅ Solicitante creado con ID: {solicitante_id}")

            # 1.b SUBIR DOCUMENTOS (si vinieron en multipart)
            documentos_creados = []
            if files_list:
                print(f"\n1️⃣b SUBIENDO DOCUMENTOS ({len(files_list)})...")
                storage = supabase.storage.from_("document")
                for idx, file in enumerate(files_list):
                    try:
                        original_name = secure_filename(file.filename or f"documento_{idx+1}")
                        ext = os.path.splitext(original_name)[1]
                        unique_name = f"{uuid.uuid4().hex}{ext}" if ext else uuid.uuid4().hex
                        storage_path = f"solicitantes/{solicitante_id}/{unique_name}"

                        content_type = getattr(file, "mimetype", None) or "application/octet-stream"
                        file_bytes = file.read()
                        storage.upload(
                            storage_path,
                            file_bytes,
                            file_options={
                                "content-type": content_type,
                                "upsert": "true",
                            },
                        )

                        public_url_resp = storage.get_public_url(storage_path)
                        public_url = None
                        if isinstance(public_url_resp, dict):
                            public_url = public_url_resp.get("publicUrl") or (
                                public_url_resp.get("data", {}) if isinstance(public_url_resp.get("data"), dict) else {}
                            ).get("publicUrl")
                        elif hasattr(public_url_resp, "data"):
                            data_dict = getattr(public_url_resp, "data", {})
                            if isinstance(data_dict, dict):
                                public_url = data_dict.get("publicUrl")
                        if not public_url:
                            public_url = str(public_url_resp)

                        doc_saved = self.documentos_model.create(
                            nombre=original_name,
                            documento_url=public_url,
                            solicitante_id=solicitante_id,
                        )
                        documentos_creados.append(doc_saved)
                        print(f"   ✅ Documento subido: {original_name}")
                    except Exception as e:
                        print(f"   ❌ Error subiendo documento {idx+1}: {e}")
                print(f"   📊 Total documentos subidos y guardados: {len(documentos_creados)}")

            # 2. CREAR UBICACIONES
            ubicaciones_creadas = []
            if datos_ubicaciones:
                print(f"\n2️⃣ CREANDO UBICACIONES...")
                for idx, ubicacion_data in enumerate(datos_ubicaciones):
                    print(f"   Ubicación {idx + 1} original: {ubicacion_data}")

                    # Procesar campos dinámicos para ubicaciones
                    detalle_direccion = ubicacion_data.get("detalle_direccion", {})

                    # Mover campos que están en la raíz pero deben ir en detalle_direccion
                    # NOTA: ciudad_residencia y departamento_residencia son campos fijos, NO se mueven
                    campos_para_mover = [
                        "direccion", "direccion_residencia", "tipo_direccion", "barrio", "estrato",
                        "telefono", "celular", "correo_personal", "recibir_correspondencia",
                        "tipo_vivienda", "paga_arriendo", "valor_mensual_arriendo", "arrendador",
                        "id_autenticacion"
                    ]
                    for campo in campos_para_mover:
                        if campo in ubicacion_data:
                            detalle_direccion[campo] = ubicacion_data.pop(campo)
                            print(f"   🔄 Movido '{campo}' a detalle_direccion")

                    # Actualizar ubicacion_data con detalle_direccion procesado
                    ubicacion_data["detalle_direccion"] = detalle_direccion

                    print(f"   Ubicación {idx + 1} procesada: {ubicacion_data}")
                    ubicacion_data["empresa_id"] = empresa_id
                    ubicacion_data["solicitante_id"] = solicitante_id

                    ubicacion_creada = self.ubicaciones_model.create(**ubicacion_data)
                    ubicaciones_creadas.append(ubicacion_creada)
                    print(f"   ✅ Ubicación {idx + 1} creada con ID: {ubicacion_creada['id']}")
            else:
                print(f"\n2️⃣ UBICACIONES: No hay datos para crear")

            # 3. CREAR ACTIVIDAD ECONÓMICA
            actividad_creada = None
            if datos_actividad:
                print(f"\n3️⃣ CREANDO ACTIVIDAD ECONÓMICA...")
                print(f"   Datos originales: {datos_actividad}")

                # Procesar campos dinámicos para actividad económica
                detalle_actividad = datos_actividad.get("detalle_actividad", {})

                # Mover campos que están en la raíz pero deben ir en detalle_actividad
                campos_para_mover = ["tipo_actividad", "tipo_actividad_economica"]
                for campo in campos_para_mover:
                    if campo in datos_actividad:
                        detalle_actividad[campo] = datos_actividad.pop(campo)
                        print(f"   🔄 Movido '{campo}' a detalle_actividad")

                # Actualizar datos_actividad con detalle_actividad procesado
                datos_actividad["detalle_actividad"] = detalle_actividad

                print(f"   Datos procesados: {datos_actividad}")
                datos_actividad["empresa_id"] = empresa_id
                datos_actividad["solicitante_id"] = solicitante_id

                actividad_creada = self.actividad_model.create(**datos_actividad)
                print(f"   ✅ Actividad económica creada con ID: {actividad_creada['id']}")
            else:
                print(f"\n3️⃣ ACTIVIDAD ECONÓMICA: No hay datos para crear")

            # 4. CREAR INFORMACIÓN FINANCIERA
            financiera_creada = None
            if datos_financiera:
                print(f"\n4️⃣ CREANDO INFORMACIÓN FINANCIERA...")
                print(f"   Datos originales: {datos_financiera}")

                # Procesar campos dinámicos para información financiera
                detalle_financiera = datos_financiera.get("detalle_financiera", {})

                # Mover campos que están en la raíz pero deben ir en detalle_financiera
                campos_para_mover = [
                    "ingreso_basico_mensual", "ingreso_variable_mensual", "otros_ingresos_mensuales",
                    "gastos_financieros_mensuales", "gastos_personales_mensuales", "ingresos_fijos_pension",
                    "ingresos_por_ventas", "ingresos_varios", "honorarios", "arriendos",
                    "ingresos_actividad_independiente", "detalle_otros_ingresos", "declara_renta"
                ]
                for campo in campos_para_mover:
                    if campo in datos_financiera:
                        detalle_financiera[campo] = datos_financiera.pop(campo)
                        print(f"   🔄 Movido '{campo}' a detalle_financiera")

                # Actualizar datos_financiera con detalle_financiera procesado
                datos_financiera["detalle_financiera"] = detalle_financiera

                # Preparar campos obligatorios (con valores por defecto si no están presentes)
                datos_para_modelo = {
                    "empresa_id": empresa_id,
                    "solicitante_id": solicitante_id,
                    "total_ingresos_mensuales": float(datos_financiera.get("total_ingresos_mensuales", 0)),
                    "total_egresos_mensuales": float(datos_financiera.get("total_egresos_mensuales", 0)),
                    "total_activos": float(datos_financiera.get("total_activos", 0)),
                    "total_pasivos": float(datos_financiera.get("total_pasivos", 0)),
                    "detalle_financiera": detalle_financiera
                }

                print(f"   Datos procesados: {datos_para_modelo}")

                financiera_creada = self.financiera_model.create(**datos_para_modelo)
                print(f"   ✅ Información financiera creada con ID: {financiera_creada['id']}")
            else:
                print(f"\n4️⃣ INFORMACIÓN FINANCIERA: No hay datos para crear")

            # 5. CREAR REFERENCIAS (JSON en una sola fila por solicitante)
            referencias_creadas = []
            if datos_referencias:
                print(f"\n5️⃣ CREANDO REFERENCIAS (JSON)...")
                for idx, referencia_data in enumerate(datos_referencias):
                    print(f"   Referencia {idx + 1} original: {referencia_data}")

                    # La referencia se almacena como objeto dentro de detalle_referencia.referencias
                    # Se admite tanto que los campos vengan en la raíz como en una subclave.
                    # Mantendremos la estructura del ejemplo del usuario.
                    nueva_ref = {}
                    # Copiar campos conocidos si vienen en la raíz
                    for k in [
                        "nombre_completo", "telefono", "celular_referencia", "relacion_referencia",
                        "ciudad", "departamento", "direccion", "snack_favorito", "si_o_no",
                        "id_tipo_referencia", "tipo_referencia"
                    ]:
                        if k in referencia_data:
                            nueva_ref[k] = referencia_data[k]

                    # Si viene un detalle_referencia.referencias[0] ya conformado, tomarlo como base
                    if isinstance(referencia_data.get("detalle_referencia"), dict):
                        arr = referencia_data["detalle_referencia"].get("referencias")
                        if isinstance(arr, list) and arr:
                            # Usar el primer elemento como base
                            base = arr[0]
                            if isinstance(base, dict):
                                base_clean = dict(base)
                                base_clean.pop("referencia_id", None)
                                nueva_ref.update(base_clean)

                    # También aceptar alias "tipo"
                    if "tipo" in referencia_data and isinstance(referencia_data["tipo"], (str, dict)):
                        nueva_ref["tipo"] = referencia_data["tipo"]
                    elif "tipo_referencia" in referencia_data and isinstance(referencia_data["tipo_referencia"], (str, dict)):
                        nueva_ref["tipo_referencia"] = referencia_data["tipo_referencia"]

                    # Limpiar vacíos y validar para evitar crear referencias "vacías"
                    cleaned = {k: v for k, v in nueva_ref.items() if v not in (None, "", [])}
                    only_tipo = set(cleaned.keys()) <= {"tipo_referencia"}

                    if not cleaned or only_tipo:
                        print(f"   ⚠️ Referencia {idx+1}: no se crea porque está vacía o solo trae 'tipo_referencia' (keys={list(nueva_ref.keys())})")
                        continue

                    # Agregar mediante el modelo JSON (generará referencia_id incremental)
                    agregado = self.referencias_model.add_referencia(
                        empresa_id=empresa_id,
                        solicitante_id=solicitante_id,
                        referencia=cleaned
                    )
                    referencias_creadas.append(agregado)
                print(f"   ✅ Total referencias agregadas: {len(referencias_creadas)}")
            else:
                print(f"\n5️⃣ REFERENCIAS: No hay datos para crear")

            # 6. CREAR SOLICITUDES
            solicitudes_creadas = []
            if datos_solicitudes:
                print(f"\n6️⃣ CREANDO SOLICITUDES...")
                for idx, solicitud_data in enumerate(datos_solicitudes):
                    print(f"   Solicitud {idx + 1}: {solicitud_data}")

                    # Extraer banco y ciudad desde campos fijos (raíz del objeto solicitud)
                    detalle_credito = solicitud_data.get("detalle_credito", {})
                    banco_nombre = solicitud_data.get("banco_nombre")  # Campo fijo en la raíz
                    ciudad = solicitud_data.get("ciudad_solicitud")    # Campo fijo en la raíz

                    print(f"   🏦 Banco extraído: {banco_nombre}")
                    print(f"   🏙️ Ciudad extraída: {ciudad}")

                    # Procesar campos anidados de tipo de crédito
                    tipo_credito = solicitud_data.get("tipo_credito")
                    if tipo_credito:
                        # Normalizar para evitar problemas por acentos/mayúsculas
                        def _norm(s: str) -> str:
                            try:
                                return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').lower()
                            except Exception:
                                return str(s).lower()

                        detalle_credito["tipo_credito"] = tipo_credito

                        credit_type_mappings = {
                            "Credito hipotecario": "credito_hipotecario",
                            "Credito de consumo": "credito_consumo",
                            "Credito comercial": "credito_comercial",
                            "Microcredito": "microcredito",
                            "Credito educativo": "credito_educativo",
                            "Credito vehicular": "credito_vehicular",
                            # Alias comunes
                            "Credito de vehiculo": "credito_vehicular",
                            "Credito de vehículo": "credito_vehicular",
                        }

                        tipo_norm = _norm(tipo_credito)
                        detected_nested = None
                        for credit_type, nested_field in credit_type_mappings.items():
                            if _norm(credit_type) in tipo_norm or (nested_field == "credito_vehicular" and "vehicul" in tipo_norm):
                                detected_nested = nested_field
                                # Solo agregar el subobjeto si no existe ya en detalle_credito
                                if nested_field in solicitud_data and nested_field not in detalle_credito:
                                    detalle_credito[nested_field] = solicitud_data[nested_field]
                                elif isinstance(detalle_credito, dict) and nested_field in detalle_credito:
                                    # Ya viene dentro de detalle_credito, mantenerlo
                                    pass
                                print(f"   📋 Procesando campos anidados para {credit_type}: {nested_field}")
                                break

                    # Copiar cualquier subobjeto de crédito conocido aunque no coincida tipo_credito
                    # Solo si no existe ya en detalle_credito para preservar la estructura original
                    known_nested_fields = [
                        "credito_hipotecario",
                        "credito_consumo",
                        "credito_comercial",
                        "microcredito",
                        "credito_educativo",
                        "credito_vehicular",
                    ]
                    for nf in known_nested_fields:
                        if nf in solicitud_data and nf not in detalle_credito:
                            detalle_credito[nf] = solicitud_data[nf]

                    # LIMPIAR campos duplicados: eliminar de la raíz de detalle_credito
                    # los campos que ya están dentro de los objetos anidados
                    for nested_field in known_nested_fields:
                        if nested_field in detalle_credito and isinstance(detalle_credito[nested_field], dict):
                            nested_obj = detalle_credito[nested_field]
                            # Eliminar de la raíz todos los campos que existen en el objeto anidado
                            campos_a_eliminar = []
                            for campo in list(detalle_credito.keys()):
                                if campo != nested_field and campo in nested_obj:
                                    campos_a_eliminar.append(campo)
                            
                            for campo in campos_a_eliminar:
                                detalle_credito.pop(campo, None)
                                print(f"   🧹 Eliminado campo duplicado '{campo}' de la raíz de detalle_credito")

                    # Extraer datos del asesor y banco desde el body principal (no desde solicitud_data)
                    nombre_asesor = body.get("nombre_asesor", "")
                    correo_asesor = body.get("correo_asesor", "")
                    nombre_banco_usuario = body.get("nombre_banco_usuario", "")
                    correo_banco_usuario = body.get("correo_banco_usuario", "")

                    print(f"   👨‍💼 Asesor: {nombre_asesor} ({correo_asesor})")
                    print(f"   🏦 Usuario banco: {nombre_banco_usuario} ({correo_banco_usuario})")

                    # Agregar estos datos al detalle_credito para que se guarden
                    if nombre_asesor:
                        detalle_credito["nombre_asesor"] = nombre_asesor
                    if correo_asesor:
                        detalle_credito["correo_asesor"] = correo_asesor
                    if nombre_banco_usuario:
                        detalle_credito["nombre_banco_usuario"] = nombre_banco_usuario
                    if correo_banco_usuario:
                        detalle_credito["correo_banco_usuario"] = correo_banco_usuario

                    # NOTA: banco_nombre y ciudad_solicitud son campos fijos, no van en detalle_credito
                    # detalle_credito solo contiene campos dinámicos

                    # Obtener user_id del header
                    user_id = request.headers.get("X-User-Id")
                    if not user_id:
                        raise ValueError("X-User-Id header es requerido para crear solicitudes")

                    # Preparar datos para el modelo (solo campos que acepta el modelo)
                    datos_para_modelo = {
                        "empresa_id": empresa_id,
                        "solicitante_id": solicitante_id,
                        "created_by_user_id": int(user_id),
                        "assigned_to_user_id": int(user_id),  # Asignar al mismo usuario que crea
                        "estado": solicitud_data.get("estado", "Pendiente"),
                        "detalle_credito": detalle_credito
                    }

                    # Agregar campos opcionales solo si están presentes
                    if banco_nombre:
                        datos_para_modelo["banco_nombre"] = banco_nombre
                    if ciudad:
                        datos_para_modelo["ciudad_solicitud"] = ciudad

                    solicitud_creada = self.solicitudes_model.create(**datos_para_modelo)
                    solicitudes_creadas.append(solicitud_creada)
                    print(f"   ✅ Solicitud {idx + 1} creada con ID: {solicitud_creada['id']}")
            else:
                print(f"\n6️⃣ SOLICITUDES: No hay datos para crear")

            # 7. PREPARAR RESPUESTA
            print(f"\n🔗 PREPARANDO RESPUESTA FINAL...")
            response_data = {
                "ok": True,
                "data": {
                    "solicitante": solicitante_creado,
                    "ubicaciones": ubicaciones_creadas,
                    "actividad_economica": actividad_creada,
                    "informacion_financiera": financiera_creada,
                    "referencias": referencias_creadas,
                    "solicitudes": solicitudes_creadas,
                    "documentos": documentos_creados,
                    "resumen": {
                        "solicitante_id": solicitante_id,
                        "total_ubicaciones": len(ubicaciones_creadas),
                        "tiene_actividad_economica": bool(actividad_creada),
                        "tiene_informacion_financiera": bool(financiera_creada),
                        "total_referencias": len(referencias_creadas),
                        "total_solicitudes": len(solicitudes_creadas)
                    }
                },
                "message": f"Registro completo creado exitosamente. Solicitante ID: {solicitante_id}"
            }

            print(f"\n📊 RESUMEN FINAL:")
            print(f"   👤 Solicitante: {solicitante_id}")
            print(f"   📍 Ubicaciones: {len(ubicaciones_creadas)}")
            print(f"   💼 Actividad económica: {'✅' if actividad_creada else '❌'}")
            print(f"   💰 Info financiera: {'✅' if financiera_creada else '❌'}")
            print(f"   👥 Referencias: {len(referencias_creadas)}")
            print(f"   📄 Solicitudes: {len(solicitudes_creadas)}")

            # 8. ENVIAR EMAILS DE CONFIRMACIÓN (SOLICITANTE, ASESOR Y BANCO)
            print(f"\n📧 ENVIANDO EMAILS DE CONFIRMACIÓN...")
            try:
                # Pasar el JSON original para extraer correos de forma robusta
                email_enviado = enviar_email_registro_completo(response_data, body)
                if email_enviado:
                    print(f"   ✅ Emails enviados exitosamente")
                    response_data["emails_enviados"] = True
                else:
                    print(f"   ⚠️ No se pudieron enviar todos los emails, pero el registro se creó correctamente")
                    response_data["emails_enviados"] = False
            except Exception as email_error:
                print(f"   ❌ Error enviando emails: {str(email_error)}")
                response_data["emails_enviados"] = False
                # No fallar la operación por error de email

            log_response(response_data)
            return jsonify(response_data), 201

        except ValueError as ve:
            log_error(ve, "ERROR DE VALIDACIÓN")
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            log_error(ex, "ERROR INESPERADO")
            return jsonify({"ok": False, "error": str(ex)}), 500

    def traer_todos_registros(self, solicitante_id: int):
        """Trae toda la información combinada de un solicitante incluyendo campos dinámicos"""
        log_request_details("TRAER TODOS REGISTROS", f"solicitante_id={solicitante_id}")

        try:
            empresa_id = self._empresa_id()

            # 1. Obtener datos del solicitante principal
            solicitante = self.model.get_by_id(id=solicitante_id, empresa_id=empresa_id)
            if not solicitante:
                return jsonify({"ok": False, "error": "Solicitante no encontrado"}), 404

            # 2. Obtener ubicaciones
            try:
                ubicaciones = self.ubicaciones_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
            except Exception as e:
                ubicaciones = []

            # 3. Obtener actividad económica
            try:
                actividad_economica_list = self.actividad_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
                actividad_economica = actividad_economica_list[0] if actividad_economica_list else None
            except Exception as e:
                actividad_economica = None

            # 4. Obtener información financiera
            try:
                financiera_list = self.financiera_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
                informacion_financiera = financiera_list[0] if financiera_list else None
            except Exception as e:
                print(f"   ❌ Error obteniendo información financiera: {e}")
                informacion_financiera = None

            # 5. Obtener referencias
            try:
                referencias = self.referencias_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
            except Exception as e:
                print(f"   ❌ Error obteniendo referencias: {e}")
                referencias = []

            # 6. Obtener solicitudes
            try:
                solicitudes = self.solicitudes_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
            except Exception as e:
                print(f"   ❌ Error obteniendo solicitudes: {e}")
                solicitudes = []

            # 7. Obtener documentos
            try:
                documentos = self.documentos_model.list(solicitante_id=solicitante_id)
            except Exception as e:
                print(f"   ❌ Error obteniendo documentos: {e}")
                documentos = []

            # 8. Combinar toda la información
            datos_completos = {
                "solicitante": solicitante,
                "ubicaciones": ubicaciones or [],
                "actividad_economica": actividad_economica,
                "informacion_financiera": informacion_financiera,
                "referencias": referencias or [],
                "solicitudes": solicitudes or [],
                "documentos": documentos or [],
                "resumen": {
                    "total_ubicaciones": len(ubicaciones) if ubicaciones else 0,
                    "tiene_actividad_economica": bool(actividad_economica),
                    "tiene_informacion_financiera": bool(informacion_financiera),
                    "total_referencias": len(referencias) if referencias else 0,
                    "total_solicitudes": len(solicitudes) if solicitudes else 0,
                    "total_documentos": len(documentos) if documentos else 0
                }
            }

            response_data = {"ok": True, "data": datos_completos}
            log_response(response_data)

            return jsonify(response_data)

        except ValueError as ve:
            log_error(ve, "ERROR DE VALIDACIÓN")
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            log_error(ex, "ERROR INESPERADO")
            return jsonify({"ok": False, "error": str(ex)}), 500

    def editar_registro_completo(self, solicitante_id: int):
        """Editar un registro completo: solicitante + ubicaciones + actividad económica + info financiera + referencias + solicitudes"""
        log_request_details("EDITAR REGISTRO COMPLETO", f"solicitante_id={solicitante_id}")

        try:
            empresa_id = self._empresa_id()
            body = request.get_json(silent=True) or {}

            print(f"\n📋 EMPRESA ID: {empresa_id}")
            print(f"📋 SOLICITANTE ID: {solicitante_id}")

            # Verificar que el solicitante existe
            solicitante_existente = self.model.get_by_id(id=solicitante_id, empresa_id=empresa_id)
            if not solicitante_existente:
                return jsonify({"ok": False, "error": "Solicitante no encontrado"}), 404

            # Extraer datos de cada entidad
            datos_solicitante = body.get("solicitante", {})
            datos_ubicaciones = body.get("ubicaciones", [])
            datos_actividad = body.get("actividad_economica", {})
            datos_financiera = body.get("informacion_financiera", {})
            datos_referencias = body.get("referencias", [])
            datos_solicitudes = body.get("solicitudes", [])

            # 1. ACTUALIZAR SOLICITANTE
            solicitante_actualizado = None
            if datos_solicitante:
                print(f"\n1️⃣ ACTUALIZANDO SOLICITANTE...")
                solicitante_actualizado = self.model.update(
                    id=solicitante_id,
                    empresa_id=empresa_id,
                    updates=datos_solicitante
                )
                print(f"   ✅ Solicitante actualizado")
            else:
                solicitante_actualizado = solicitante_existente

            # 2. ACTUALIZAR UBICACIONES
            ubicaciones_actualizadas = []
            if datos_ubicaciones:
                print(f"\n2️⃣ ACTUALIZANDO UBICACIONES...")

                # Obtener ubicaciones existentes del solicitante
                ubicaciones_existentes = self.ubicaciones_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
                print(f"   📍 Ubicaciones existentes encontradas: {len(ubicaciones_existentes)}")

                for idx, ubicacion_data in enumerate(datos_ubicaciones):
                    # Procesar campos dinámicos
                    detalle_direccion = ubicacion_data.get("detalle_direccion", {})
                    campos_para_mover = [
                        "direccion", "direccion_residencia", "tipo_direccion", "barrio", "estrato",
                        "telefono", "celular", "correo_personal", "recibir_correspondencia",
                        "tipo_vivienda", "paga_arriendo", "valor_mensual_arriendo", "arrendador"
                    ]
                    for campo in campos_para_mover:
                        if campo in ubicacion_data:
                            detalle_direccion[campo] = ubicacion_data.pop(campo)

                    ubicacion_data["detalle_direccion"] = detalle_direccion

                    # Buscar ubicación existente para actualizar
                    ubicacion_id = ubicacion_data.get("id")
                    ubicacion_existente = None

                    # Si hay ID específico, buscar por ID
                    if ubicacion_id and str(ubicacion_id).strip() and str(ubicacion_id) != "0":
                        ubicacion_existente = next((u for u in ubicaciones_existentes if u["id"] == int(ubicacion_id)), None)
                        print(f"   🔍 Buscando ubicación por ID {ubicacion_id}: {'Encontrada' if ubicacion_existente else 'No encontrada'}")

                    # Si no hay ID o no se encontró, usar ubicación existente por índice
                    if not ubicacion_existente and idx < len(ubicaciones_existentes):
                        ubicacion_existente = ubicaciones_existentes[idx]
                        print(f"   🔄 Usando ubicación existente por índice {idx}, ID: {ubicacion_existente['id']}")

                    # Actualizar o crear según corresponda
                    if ubicacion_existente:
                        print(f"   ✏️ ACTUALIZANDO ubicación ID: {ubicacion_existente['id']}")
                        ubicacion_actualizada = self.ubicaciones_model.update(
                            id=ubicacion_existente["id"],
                            empresa_id=empresa_id,
                            updates={k: v for k, v in ubicacion_data.items() if k != "id"}
                        )
                        print(f"   ✅ Ubicación {ubicacion_existente['id']} actualizada exitosamente")
                    else:
                        print(f"   🆕 CREANDO nueva ubicación (no hay ubicaciones existentes para índice {idx})")
                        ubicacion_data["empresa_id"] = empresa_id
                        ubicacion_data["solicitante_id"] = solicitante_id
                        ubicacion_actualizada = self.ubicaciones_model.create(**ubicacion_data)

                    ubicaciones_actualizadas.append(ubicacion_actualizada)

            # 3. ACTUALIZAR ACTIVIDAD ECONÓMICA
            actividad_actualizada = None
            if datos_actividad:
                print(f"\n3️⃣ ACTUALIZANDO ACTIVIDAD ECONÓMICA...")
                detalle_actividad = datos_actividad.get("detalle_actividad", {})
                campos_para_mover = ["tipo_actividad", "tipo_actividad_economica"]
                for campo in campos_para_mover:
                    if campo in datos_actividad:
                        detalle_actividad[campo] = datos_actividad.pop(campo)

                datos_actividad["detalle_actividad"] = detalle_actividad

                # Buscar actividad existente
                actividades_existentes = self.actividad_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
                if actividades_existentes and len(actividades_existentes) > 0:
                    actividad_id = actividades_existentes[0]["id"]
                    actividad_actualizada = self.actividad_model.update(
                        id=actividad_id,
                        empresa_id=empresa_id,
                        updates=datos_actividad
                    )
                else:
                    datos_actividad["empresa_id"] = empresa_id
                    datos_actividad["solicitante_id"] = solicitante_id
                    actividad_actualizada = self.actividad_model.create(**datos_actividad)

            # 4. ACTUALIZAR INFORMACIÓN FINANCIERA
            financiera_actualizada = None
            if datos_financiera:
                print(f"\n4️⃣ ACTUALIZANDO INFORMACIÓN FINANCIERA...")
                detalle_financiera = datos_financiera.get("detalle_financiera", {})
                campos_para_mover = [
                    "ingreso_basico_mensual", "ingreso_variable_mensual", "otros_ingresos_mensuales",
                    "gastos_financieros_mensuales", "gastos_personales_mensuales", "declara_renta"
                ]
                for campo in campos_para_mover:
                    if campo in datos_financiera:
                        detalle_financiera[campo] = datos_financiera.pop(campo)

                datos_financiera["detalle_financiera"] = detalle_financiera

                # Buscar información financiera existente
                financieras_existentes = self.financiera_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
                if financieras_existentes and len(financieras_existentes) > 0:
                    financiera_id = financieras_existentes[0]["id"]
                    financiera_actualizada = self.financiera_model.update(
                        id=financiera_id,
                        empresa_id=empresa_id,
                        updates=datos_financiera
                    )
                else:
                    # Crear nueva - preparar campos obligatorios
                    datos_para_modelo = {
                        "empresa_id": empresa_id,
                        "solicitante_id": solicitante_id,
                        "total_ingresos_mensuales": float(datos_financiera.get("total_ingresos_mensuales", 0)),
                        "total_egresos_mensuales": float(datos_financiera.get("total_egresos_mensuales", 0)),
                        "total_activos": float(datos_financiera.get("total_activos", 0)),
                        "total_pasivos": float(datos_financiera.get("total_pasivos", 0)),
                        "detalle_financiera": detalle_financiera
                    }
                    financiera_actualizada = self.financiera_model.create(**datos_para_modelo)

            # 5. ACTUALIZAR REFERENCIAS (JSON en una sola fila por solicitante)
            referencias_actualizadas = []
            if datos_referencias:
                print(f"\n5️⃣ ACTUALIZANDO REFERENCIAS (JSON)...")
                for idx, referencia_data in enumerate(datos_referencias):
                    # Si viene referencia_id (o alias id), actualizamos campos; si no, agregamos como nueva
                    ref_id = referencia_data.get("referencia_id") or referencia_data.get("id")

                    # Construir payload de campos a actualizar/agregar desde la raíz o desde detalle
                    payload = {}
                    for k in [
                        "nombre_completo", "telefono", "celular_referencia", "relacion_referencia",
                        "ciudad", "departamento", "direccion", "snack_favorito", "si_o_no",
                        "id_tipo_referencia", "tipo_referencia"
                    ]:
                        if k in referencia_data:
                            payload[k] = referencia_data[k]

                    if isinstance(referencia_data.get("detalle_referencia"), dict):
                        arr = referencia_data["detalle_referencia"].get("referencias")
                        if isinstance(arr, list) and arr:
                            base = arr[0]
                            if isinstance(base, dict):
                                base_clean = dict(base)
                                base_clean.pop("referencia_id", None)
                                payload.update(base_clean)

                    # tipo/tipo_referencia como alias
                    if "tipo" in referencia_data and isinstance(referencia_data["tipo"], (str, dict)):
                        payload["tipo"] = referencia_data["tipo"]
                    elif "tipo_referencia" in referencia_data and isinstance(referencia_data["tipo_referencia"], (str, dict)):
                        payload["tipo_referencia"] = referencia_data["tipo_referencia"]

                    # Limpiar vacíos y validar para evitar crear referencias "vacías"
                    cleaned = {k: v for k, v in payload.items() if v not in (None, "", [])}
                    only_tipo = set(cleaned.keys()) <= {"tipo_referencia"}

                    if ref_id is not None and str(ref_id).strip() != "":
                        if not cleaned:
                            print(f"   ⚠️ Referencia {ref_id} con payload vacío tras limpiar, se omite actualización")
                            continue
                        actualizado = self.referencias_model.update_referencia_fields(
                            empresa_id=empresa_id,
                            solicitante_id=solicitante_id,
                            referencia_id=int(ref_id),
                            updates=cleaned
                        )
                        if actualizado is not None:
                            referencias_actualizadas.append(actualizado)
                        else:
                            print(f"   ❌ Referencia no encontrada para actualizar. referencia_id solicitado: {ref_id}")
                    else:
                        if only_tipo:
                            print(f"   ⚠️ Referencia {idx+1}: no se crea porque solo trae 'tipo_referencia' sin campos informativos")
                            continue
                        agregado = self.referencias_model.add_referencia(
                            empresa_id=empresa_id,
                            solicitante_id=solicitante_id,
                            referencia=cleaned
                        )
                        referencias_actualizadas.append(agregado)

            # 6. ACTUALIZAR SOLICITUDES
            solicitudes_actualizadas = []

            if datos_solicitudes:
                print(f"\n6️⃣ ACTUALIZANDO SOLICITUDES...")
                user_id = request.headers.get("X-User-Id")
                if not user_id:
                    raise ValueError("X-User-Id header es requerido para actualizar solicitudes")

                # Obtener solicitudes existentes del solicitante
                solicitudes_existentes = self.solicitudes_model.list(empresa_id=empresa_id, solicitante_id=solicitante_id)
                print(f"   📋 Solicitudes existentes encontradas: {len(solicitudes_existentes)}")

                for idx, solicitud_data in enumerate(datos_solicitudes):
                    detalle_credito = solicitud_data.get("detalle_credito", {})

                    # Procesar campos anidados de tipo de crédito
                    tipo_credito = solicitud_data.get("tipo_credito")
                    if tipo_credito:
                        # Normalizar para evitar problemas por acentos/mayúsculas
                        def _norm(s: str) -> str:
                            try:
                                return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').lower()
                            except Exception:
                                return str(s).lower()

                        detalle_credito["tipo_credito"] = tipo_credito

                        credit_type_mappings = {
                            "Credito hipotecario": "credito_hipotecario",
                            "Credito de consumo": "credito_consumo",
                            "Credito comercial": "credito_comercial",
                            "Microcredito": "microcredito",
                            "Credito educativo": "credito_educativo",
                            "Credito vehicular": "credito_vehicular",
                            # Alias comunes
                            "Credito de vehiculo": "credito_vehicular",
                            "Credito de vehículo": "credito_vehicular",
                        }

                        tipo_norm = _norm(tipo_credito)
                        detected_nested = None
                        for credit_type, nested_field in credit_type_mappings.items():
                            if _norm(credit_type) in tipo_norm or (nested_field == "credito_vehicular" and "vehicul" in tipo_norm):
                                detected_nested = nested_field
                                # Solo agregar el subobjeto si no existe ya en detalle_credito
                                if nested_field in solicitud_data and nested_field not in detalle_credito:
                                    detalle_credito[nested_field] = solicitud_data[nested_field]
                                elif isinstance(detalle_credito, dict) and nested_field in detalle_credito:
                                    # Ya viene dentro de detalle_credito, mantenerlo
                                    pass
                                print(f"   📋 Procesando campos anidados para {credit_type}: {nested_field}")
                                break

                    # Copiar cualquier subobjeto de crédito conocido aunque no coincida tipo_credito
                    # Solo si no existe ya en detalle_credito para preservar la estructura original
                    known_nested_fields = [
                        "credito_hipotecario",
                        "credito_consumo",
                        "credito_comercial",
                        "microcredito",
                        "credito_educativo",
                        "credito_vehicular",
                    ]
                    for nf in known_nested_fields:
                        if nf in solicitud_data and nf not in detalle_credito:
                            detalle_credito[nf] = solicitud_data[nf]

                    # LIMPIAR campos duplicados: eliminar de la raíz de detalle_credito
                    # los campos que ya están dentro de los objetos anidados
                    for nested_field in known_nested_fields:
                        if nested_field in detalle_credito and isinstance(detalle_credito[nested_field], dict):
                            nested_obj = detalle_credito[nested_field]
                            # Eliminar de la raíz todos los campos que existen en el objeto anidado
                            campos_a_eliminar = []
                            for campo in list(detalle_credito.keys()):
                                if campo != nested_field and campo in nested_obj:
                                    campos_a_eliminar.append(campo)
                            
                            for campo in campos_a_eliminar:
                                detalle_credito.pop(campo, None)
                                print(f"   🧹 Eliminado campo duplicado '{campo}' de la raíz de detalle_credito")

                    # Extraer datos del asesor y banco desde el body principal (no desde solicitud_data)
                    nombre_asesor = body.get("nombre_asesor", "")
                    correo_asesor = body.get("correo_asesor", "")
                    nombre_banco_usuario = body.get("nombre_banco_usuario", "")
                    correo_banco_usuario = body.get("correo_banco_usuario", "")

                    print(f"   👨‍💼 Asesor: {nombre_asesor} ({correo_asesor})")
                    print(f"   🏦 Usuario banco: {nombre_banco_usuario} ({correo_banco_usuario})")

                    # Agregar estos datos al detalle_credito para que se guarden
                    if nombre_asesor:
                        detalle_credito["nombre_asesor"] = nombre_asesor
                    if correo_asesor:
                        detalle_credito["correo_asesor"] = correo_asesor
                    if nombre_banco_usuario:
                        detalle_credito["nombre_banco_usuario"] = nombre_banco_usuario
                    if correo_banco_usuario:
                        detalle_credito["correo_banco_usuario"] = correo_banco_usuario

                    datos_para_modelo = {
                        "estado": solicitud_data.get("estado", "Pendiente"),
                        "detalle_credito": detalle_credito
                    }

                    if solicitud_data.get("banco_nombre"):
                        datos_para_modelo["banco_nombre"] = solicitud_data["banco_nombre"]
                    if solicitud_data.get("ciudad_solicitud"):
                        datos_para_modelo["ciudad_solicitud"] = solicitud_data["ciudad_solicitud"]

                    # Buscar solicitud existente para actualizar
                    solicitud_id = solicitud_data.get("id")
                    solicitud_existente = None

                    # Si hay ID específico, buscar por ID
                    if solicitud_id and str(solicitud_id).strip() and str(solicitud_id) != "0":
                        solicitud_existente = next((s for s in solicitudes_existentes if s["id"] == int(solicitud_id)), None)
                        print(f"   🔍 Buscando por ID {solicitud_id}: {'Encontrada' if solicitud_existente else 'No encontrada'}")

                    # Si no hay ID o no se encontró, usar la primera solicitud existente
                    if not solicitud_existente and solicitudes_existentes:
                        solicitud_existente = solicitudes_existentes[0]
                        print(f"   🔄 Usando primera solicitud existente ID: {solicitud_existente['id']}")


                    # Actualizar o crear según corresponda
                    if solicitud_existente:
                        print(f"   ✏️ ACTUALIZANDO solicitud ID: {solicitud_existente['id']}")

                        # Separar detalle_credito de otros campos base
                        detalle_credito = datos_para_modelo.pop("detalle_credito", {})
                        base_updates = datos_para_modelo if datos_para_modelo else None

                        solicitud_actualizada = self.solicitudes_model.update(
                            id=solicitud_existente["id"],
                            empresa_id=empresa_id,
                            base_updates=base_updates,
                            detalle_credito_merge=detalle_credito
                        )
                        print(f"   ✅ Solicitud {solicitud_existente['id']} actualizada exitosamente")
                    else:
                        print(f"   🆕 CREANDO nueva solicitud (no hay solicitudes existentes)")
                        datos_para_modelo.update({
                            "empresa_id": empresa_id,
                            "solicitante_id": solicitante_id,
                            "created_by_user_id": int(user_id),
                            "assigned_to_user_id": int(user_id)
                        })
                        solicitud_actualizada = self.solicitudes_model.create(**datos_para_modelo)

                    solicitudes_actualizadas.append(solicitud_actualizada)

            # 7. ENVIAR EMAILS DE CONFIRMACIÓN (SOLICITANTE, ASESOR Y BANCO)
            print(f"\n📧 ENVIANDO EMAILS DE CONFIRMACIÓN...")
            try:
                # Preparar response_data temporal para el email
                response_data_temp = {
                    "ok": True,
                    "data": {
                        "solicitante": solicitante_actualizado,
                        "ubicaciones": ubicaciones_actualizadas,
                        "actividad_economica": actividad_actualizada,
                        "informacion_financiera": financiera_actualizada,
                        "referencias": referencias_actualizadas,
                        "solicitudes": solicitudes_actualizadas
                    }
                }

                # Pasar el JSON original para extraer correos de forma robusta
                email_enviado = enviar_email_registro_completo(response_data_temp, body)
                emails_enviados_exitosamente = email_enviado
                if email_enviado:
                    print(f"   ✅ Emails enviados exitosamente")
                else:
                    print(f"   ⚠️ No se pudieron enviar todos los emails, pero la actualización se completó correctamente")
            except Exception as email_error:
                print(f"   ❌ Error enviando emails: {str(email_error)}")
                emails_enviados_exitosamente = False
                # No fallar la operación por error de email

            # 8. PREPARAR RESPUESTA
            response_data = {
                "ok": True,
                "data": {
                    "solicitante": solicitante_actualizado,
                    "ubicaciones": ubicaciones_actualizadas,
                    "actividad_economica": actividad_actualizada,
                    "informacion_financiera": financiera_actualizada,
                    "referencias": referencias_actualizadas,
                    "solicitudes": solicitudes_actualizadas,
                    "resumen": {
                        "solicitante_id": solicitante_id,
                        "total_ubicaciones": len(ubicaciones_actualizadas),
                        "tiene_actividad_economica": bool(actividad_actualizada),
                        "tiene_informacion_financiera": bool(financiera_actualizada),
                        "total_referencias": len(referencias_actualizadas),
                        "total_solicitudes": len(solicitudes_actualizadas)
                    }
                },
                "message": f"Registro completo actualizado exitosamente. Solicitante ID: {solicitante_id}",
                "emails_enviados": emails_enviados_exitosamente  # Indicar si se enviaron emails
            }

            log_response(response_data)
            return jsonify(response_data), 200

        except ValueError as ve:
            log_error(ve, "ERROR DE VALIDACIÓN")
            return jsonify({"ok": False, "error": str(ve)}), 400
        except Exception as ex:
            log_error(ex, "ERROR INESPERADO")
            return jsonify({"ok": False, "error": str(ex)}), 500
